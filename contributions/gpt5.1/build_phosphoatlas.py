#!/usr/bin/env python3
"""
Build a comprehensive human kinase–substrate–phosphosite atlas
from locally available databases outside agent-pa-benchmark.
"""
import csv
import json
import re
from pathlib import Path

import openpyxl

# Output location (this folder)
BASE = Path(__file__).resolve().parent
OUTPUT_XLSX = BASE / "phosphoatlas_kinase_substrate_sites.xlsx"
OUTPUT_JSON = BASE / "phosphoatlas_kinase_substrate_sites.json"

# Input sources (outside agent-pa-benchmark)
WORKSPACE = Path("/Users/lukasamare/.openclaw/workspace")
UNIPROT_ACC_GENE = WORKSPACE / "data/atlas_build/uniprot_acc_gene.tsv"
PHOSPHOELM_DUMP = WORKSPACE / "data/atlas_build/phosphoelm_dump/phosphoELM_all_2015-04.dump"
PHOSPHOSIGNOR = WORKSPACE / "tmp_pa_live/phosphosignor_kinaseALL.tsv"

# Header layout based on gold_standard/sample_PA2.xlsx
HEADERS = [
    "KINASE GENE",
    "KINASE common name",
    "KIN_ACC_ID",
    "SUBSTRATE common name",
    "SUB_GENE_ID",
    "SUB_ACC_ID",
    "SUBSTRATE GENE",
    "SUB_MOD_RSD",
    "SITE_GRP_ID",
    "SITE_+/-7_AA",
    "BE AWARE: available in PA2_2023, or in PA1_2016_HTKAM2_only",
    "",  # trailing empty column in sample
]


def load_uniprot_gene_map(path: Path) -> dict:
    acc_to_gene = {}
    with path.open() as f:
        r = csv.DictReader(f, delimiter="\t")
        for row in r:
            acc = row.get("Entry")
            gene = row.get("Gene Names (primary)")
            if acc and gene:
                acc_to_gene[acc] = gene
    return acc_to_gene


def heptamer_from_sequence(seq: str, pos: int) -> str:
    if not seq or not pos:
        return ""
    idx = pos - 1
    start = max(0, idx - 3)
    end = min(len(seq), idx + 4)
    return seq[start:end]


def parse_phosphosignor(path: Path, acc_to_gene: dict) -> list:
    records = []
    with path.open() as f:
        r = csv.DictReader(f, delimiter="\t")
        for row in r:
            if row.get("mechanism") and "phosphorylation" not in row["mechanism"].lower():
                continue
            kinase_acc = row.get("entityA", "").strip()
            kinase_name = row.get("entityA_name", "").strip()

            substrate_entity = row.get("entityB", "").strip()
            substrate_name = row.get("entityB_name", "").strip()

            # Parse substrate accession and phosphosite from entityB like P19419_phSer383
            substrate_acc = ""
            site = ""
            if "_ph" in substrate_entity:
                acc_part, ph_part = substrate_entity.split("_ph", 1)
                substrate_acc = acc_part
                m = re.match(r"(Ser|Thr|Tyr)(\d+)", ph_part)
                if m:
                    aa = {"Ser": "S", "Thr": "T", "Tyr": "Y"}[m.group(1)]
                    site = f"{aa}{m.group(2)}"

            # Fallback substrate accession from name if looks like accession
            if not substrate_acc and re.match(r"^[A-Z0-9]{6,10}$", substrate_entity):
                substrate_acc = substrate_entity

            # Determine gene symbols
            kinase_gene = acc_to_gene.get(kinase_acc) or kinase_name
            substrate_gene = acc_to_gene.get(substrate_acc)
            if not substrate_gene and "_ph" in substrate_name:
                substrate_gene = substrate_name.split("_ph", 1)[0]

            rec = {
                "KINASE GENE": kinase_gene or "",
                "KINASE common name": "",
                "KIN_ACC_ID": kinase_acc if re.match(r"^[A-Z0-9]{6,10}$", kinase_acc) else "",
                "SUBSTRATE common name": "",
                "SUB_GENE_ID": "",
                "SUB_ACC_ID": substrate_acc or "",
                "SUBSTRATE GENE": substrate_gene or "",
                "SUB_MOD_RSD": site or "",
                "SITE_GRP_ID": "",
                "SITE_+/-7_AA": "",
                "BE AWARE: available in PA2_2023, or in PA1_2016_HTKAM2_only": "SIGNOR",
                "": "",
            }
            # Only include if we have a site and kinase/substrate identifiers
            if rec["SUB_MOD_RSD"] and (rec["KINASE GENE"] or rec["KIN_ACC_ID"]):
                records.append(rec)
    return records


def parse_phosphoelm(path: Path, acc_to_gene: dict) -> list:
    records = []
    with path.open() as f:
        r = csv.DictReader(f, delimiter="\t")
        for row in r:
            if row.get("species") != "Homo sapiens":
                continue
            if not row.get("kinases"):
                continue
            acc = row.get("acc", "").strip()
            seq = row.get("sequence", "")
            try:
                pos = int(row.get("position") or 0)
            except ValueError:
                pos = 0
            code = row.get("code", "").strip()
            site = f"{code}{pos}" if code and pos else ""
            hept = heptamer_from_sequence(seq, pos)
            substrate_gene = acc_to_gene.get(acc, "")

            # Split kinases (common delimiter is '/'; also support ';' and ',')
            kinases = re.split(r"[\/;,]", row.get("kinases", ""))
            for kinase in [k.strip() for k in kinases if k.strip()]:
                rec = {
                    "KINASE GENE": kinase,
                    "KINASE common name": "",
                    "KIN_ACC_ID": "",
                    "SUBSTRATE common name": "",
                    "SUB_GENE_ID": "",
                    "SUB_ACC_ID": acc,
                    "SUBSTRATE GENE": substrate_gene,
                    "SUB_MOD_RSD": site,
                    "SITE_GRP_ID": "",
                    "SITE_+/-7_AA": hept,
                    "BE AWARE: available in PA2_2023, or in PA1_2016_HTKAM2_only": "PhosphoELM",
                    "": "",
                }
                if rec["SUB_MOD_RSD"] and rec["KINASE GENE"]:
                    records.append(rec)
    return records


def merge_records(records: list) -> list:
    merged = {}
    for rec in records:
        key = (
            rec.get("KINASE GENE", ""),
            rec.get("KIN_ACC_ID", ""),
            rec.get("SUBSTRATE GENE", ""),
            rec.get("SUB_ACC_ID", ""),
            rec.get("SUB_MOD_RSD", ""),
        )
        if key not in merged:
            merged[key] = rec
        else:
            # merge sources
            src = merged[key].get(
                "BE AWARE: available in PA2_2023, or in PA1_2016_HTKAM2_only", ""
            )
            new_src = rec.get(
                "BE AWARE: available in PA2_2023, or in PA1_2016_HTKAM2_only", ""
            )
            if new_src and new_src not in src:
                merged[key][
                    "BE AWARE: available in PA2_2023, or in PA1_2016_HTKAM2_only"
                ] = (src + "; " + new_src) if src else new_src
            # fill missing fields
            for h in HEADERS:
                if not merged[key].get(h) and rec.get(h):
                    merged[key][h] = rec[h]
    return list(merged.values())


def write_outputs(records: list):
    # XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "phosphoatlas"
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for r_idx, rec in enumerate(records, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=rec.get(header))
    wb.save(OUTPUT_XLSX)

    # JSON
    with OUTPUT_JSON.open("w") as f:
        json.dump(records, f, ensure_ascii=False)


if __name__ == "__main__":
    acc_to_gene = load_uniprot_gene_map(UNIPROT_ACC_GENE)
    records = []
    records.extend(parse_phosphosignor(PHOSPHOSIGNOR, acc_to_gene))
    records.extend(parse_phosphoelm(PHOSPHOELM_DUMP, acc_to_gene))
    merged = merge_records(records)
    write_outputs(merged)
    print(f"Wrote {OUTPUT_XLSX} and {OUTPUT_JSON} with {len(merged)} rows")
