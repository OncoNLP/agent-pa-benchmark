#!/usr/bin/env python3
"""
Step 1: Parse PhosphoAtlas 2.0 XLSX into a structured testing dataset.

Uses sample_PA2.xlsx as the FORMAT REFERENCE for column mapping.
Accepts ANY PhosphoAtlas spreadsheet as input (PA1, PA2, combined).

PA2 sample format columns (from sample_PA2.xlsx):
  A: KINASE GENE
  B: KINASE common name
  C: KIN_ACC_ID          (kinase UniProt accession)
  D: SUBSTRATE common name
  E: SUB_GENE_ID         (NCBI Entrez Gene ID)
  F: SUB_ACC_ID          (substrate UniProt accession)
  G: SUBSTRATE GENE
  H: SUB_MOD_RSD         (phosphorylation site, e.g. Y226)
  I: SITE_GRP_ID         (PhosphoSitePlus group ID)
  J: SITE_+/-7_AA        (heptameric peptide around phospho-site)
  K: version/source tag  (e.g. "PA2_2023 and PA1_2016")

Usage:
  python -m src.step1_parse_gold.parse_pa2 \
      --input gold_standard/input/PhosphoAtlas2.xlsx \
      --output gold_standard/parsed/phosphoatlas_gold.json

  # Or with multiple sheets:
  python -m src.step1_parse_gold.parse_pa2 \
      --input gold_standard/input/PA_combined.xlsx \
      --sheets "PA2_2023" "PA1_2016"
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


# Column name variants we accept (case-insensitive matching)
COLUMN_MAP = {
    "kinase_gene": ["kinase gene", "kinase_gene", "kin_gene"],
    "kinase_name": ["kinase common name", "kinase_name", "kin_name"],
    "kinase_uniprot": ["kin_acc_id", "kinase_acc_id", "kinase_uniprot"],
    "substrate_name": ["substrate common name", "substrate_name", "sub_name"],
    "substrate_gene_id": ["sub_gene_id", "substrate_gene_id"],
    "substrate_uniprot": ["sub_acc_id", "substrate_acc_id", "sub_uniprot"],
    "substrate_gene": ["substrate gene", "substrate_gene", "sub_gene"],
    "phospho_site": ["sub_mod_rsd", "phospho_site", "site"],
    "site_group_id": ["site_grp_id", "site_group_id"],
    "heptameric_peptide": ["site_+/-7_aa", "site_7_aa", "heptameric_peptide", "peptide"],
    "pa_version": [
        "be aware",
        "be aware: available in pa2_2023, or in pa1_2016_htkam2_only",
        "version",
        "pa_version",
        "source",
    ],
}


def match_column(header: str):
    """Match a header string to a canonical column name.

    Finds the BEST (longest matching variant) to avoid ambiguity
    (e.g., 'site_grp_id' must match site_group_id, not phospho_site's 'site').
    """
    h = header.strip().lower()
    best_canonical = None
    best_len = 0
    for canonical, variants in COLUMN_MAP.items():
        for v in variants:
            if h == v:
                return canonical  # exact match wins immediately
            if (v in h or h in v) and len(v) > best_len:
                best_canonical = canonical
                best_len = len(v)
    return best_canonical


def parse_sheet(ws, sheet_name: str = "") -> list[dict]:
    """Parse a single worksheet into a list of entry dicts.

    Uses row iteration (ws.iter_rows or ws.rows) for performance with
    openpyxl read_only mode. Cell-by-cell access via ws.cell() is O(n^2)
    in read_only mode and will hang on large files.
    """
    # Build col_mapping: canonical_name -> 0-based column index
    # by iterating the first row
    col_mapping = {}
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        header_row = row
        break

    if header_row is None:
        print(f"  WARNING: Sheet '{sheet_name}' is empty. Skipping.")
        return []

    for i, cell in enumerate(header_row):
        h = cell.value
        if not h:
            continue
        # Only scan first 20 columns
        if i >= 20:
            break
        canon = match_column(str(h))
        if canon and canon not in col_mapping:
            col_mapping[canon] = i

    if "kinase_gene" not in col_mapping or "substrate_gene" not in col_mapping:
        headers_raw = [c.value for c in header_row[:20]]
        print(f"  WARNING: Sheet '{sheet_name}' missing KINASE GENE or SUBSTRATE GENE columns. Skipping.")
        print(f"  Found headers: {headers_raw}")
        return []

    # Helper to safely get string value from a row tuple by canonical name
    def val(row_cells, canonical):
        idx = col_mapping.get(canonical)
        if idx is None or idx >= len(row_cells):
            return ""
        v = row_cells[idx]
        return str(v).strip() if v is not None else ""

    def raw(row_cells, canonical):
        idx = col_mapping.get(canonical)
        if idx is None or idx >= len(row_cells):
            return None
        return row_cells[idx]

    entries = []
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if row_count % 5000 == 0:
            print(f"    ...{row_count} rows processed")

        kinase_gene = val(row, "kinase_gene")
        substrate_gene = val(row, "substrate_gene")
        phospho_site = val(row, "phospho_site")

        if not kinase_gene or not substrate_gene or not phospho_site:
            continue

        entry = {
            "kinase_gene": kinase_gene,
            "kinase_name": val(row, "kinase_name"),
            "kinase_uniprot": val(row, "kinase_uniprot"),
            "substrate_name": val(row, "substrate_name"),
            "substrate_gene_id": val(row, "substrate_gene_id"),
            "substrate_uniprot": val(row, "substrate_uniprot"),
            "substrate_gene": substrate_gene,
            "phospho_site": phospho_site,
            "site_group_id": raw(row, "site_group_id"),
            "heptameric_peptide": val(row, "heptameric_peptide"),
            "pa_version": val(row, "pa_version"),
            "source_sheet": sheet_name,
        }

        # Clean up: remove semicolons from gene IDs (PA format quirk)
        entry["substrate_gene_id"] = entry["substrate_gene_id"].rstrip(";").strip()

        # Normalize site_group_id to int or None
        if entry["site_group_id"] is not None:
            try:
                entry["site_group_id"] = int(entry["site_group_id"])
            except (ValueError, TypeError):
                entry["site_group_id"] = None

        entries.append(entry)

    return entries


def filter_pa2_only(entries: list[dict]) -> list[dict]:
    """Filter to PA2_2023 entries only (exclude PA1-only entries)."""
    pa2_entries = []
    for e in entries:
        version = e.get("pa_version", "").lower()
        # Keep if version contains "pa2" or is empty (assume PA2)
        if "pa2" in version or not version or "htkam2" not in version:
            pa2_entries.append(e)
    return pa2_entries


def deduplicate(entries: list[dict]) -> list[dict]:
    """Deduplicate by (kinase_gene, substrate_gene, phospho_site) triplet.

    When duplicates exist, prefer entries with more complete data.
    """
    seen = {}
    for e in entries:
        key = f"{e['kinase_gene']}|{e['substrate_gene']}|{e['phospho_site']}"
        if key not in seen:
            seen[key] = e
        else:
            existing = seen[key]
            # Prefer entries with more filled fields
            new_score = sum(1 for v in e.values() if v)
            old_score = sum(1 for v in existing.values() if v)
            if new_score > old_score:
                seen[key] = e
    return list(seen.values())


def build_gold_standard(entries: list[dict]) -> dict:
    """Build the gold standard JSON structure indexed by kinase."""
    by_kinase = defaultdict(list)
    for e in entries:
        by_kinase[e["kinase_gene"]].append(e)

    kinases = {}
    for kinase_gene in sorted(by_kinase.keys()):
        kinase_entries = sorted(
            by_kinase[kinase_gene],
            key=lambda x: (x["substrate_gene"], x["phospho_site"]),
        )
        kinases[kinase_gene] = {
            "kinase_gene": kinase_gene,
            "entry_count": len(kinase_entries),
            "entries": kinase_entries,
        }

    unique_substrates = set()
    for e in entries:
        unique_substrates.add(e["substrate_gene"])

    return {
        "metadata": {
            "source": "PhosphoAtlas (Olow et al., Cancer Research 2016)",
            "dataset_version": "PA2_2023 (deduplicated)",
            "total_entries": len(entries),
            "unique_kinases": len(kinases),
            "unique_substrates": len(unique_substrates),
            "unique_triplets": len(entries),
        },
        "kinases": kinases,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Parse PhosphoAtlas XLSX into gold standard JSON for benchmarking."
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to PhosphoAtlas XLSX file (any version)."
    )
    parser.add_argument(
        "--output", default="gold_standard/parsed/phosphoatlas_gold.json",
        help="Output JSON path (default: gold_standard/parsed/phosphoatlas_gold.json)."
    )
    parser.add_argument(
        "--sheets", nargs="*", default=None,
        help="Sheet names to parse (default: all sheets)."
    )
    parser.add_argument(
        "--pa2-only", action="store_true",
        help="Filter to PA2_2023 entries only (exclude PA1-only)."
    )
    parser.add_argument(
        "--sample-format", default="gold_standard/sample_PA2.xlsx",
        help="Path to sample_PA2.xlsx for format verification."
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        print(f"Place your PhosphoAtlas XLSX in: gold_standard/input/")
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Verify sample format exists
    sample_path = Path(args.sample_format)
    if sample_path.exists():
        print(f"Format reference: {sample_path}")

    # Parse XLSX
    print(f"Parsing: {input_path}")
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    sheets_to_parse = args.sheets if args.sheets else wb.sheetnames
    print(f"Sheets: {sheets_to_parse}")

    all_entries = []
    for sheet_name in sheets_to_parse:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found. Skipping.")
            continue
        ws = wb[sheet_name]
        entries = parse_sheet(ws, sheet_name)
        print(f"  Sheet '{sheet_name}': {len(entries)} entries parsed")
        all_entries.extend(entries)

    wb.close()

    if not all_entries:
        print("ERROR: No entries parsed. Check your XLSX format against sample_PA2.xlsx.")
        sys.exit(1)

    # Filter PA2-only if requested
    if args.pa2_only:
        before = len(all_entries)
        all_entries = filter_pa2_only(all_entries)
        print(f"Filtered PA2-only: {before} -> {len(all_entries)} entries")

    # Deduplicate
    before = len(all_entries)
    all_entries = deduplicate(all_entries)
    print(f"Deduplicated: {before} -> {len(all_entries)} unique triplets")

    # Build gold standard
    gold = build_gold_standard(all_entries)

    # Save
    with open(output_path, "w") as f:
        json.dump(gold, f, indent=2)

    print(f"\nGold standard saved: {output_path}")
    print(f"  Total entries: {gold['metadata']['total_entries']}")
    print(f"  Unique kinases: {gold['metadata']['unique_kinases']}")
    print(f"  Unique substrates: {gold['metadata']['unique_substrates']}")

    # Also generate a flat CSV for easy inspection
    csv_path = output_path.with_suffix(".csv")
    import csv
    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "kinase_gene", "kinase_name", "kinase_uniprot",
            "substrate_gene", "substrate_name", "substrate_uniprot",
            "substrate_gene_id", "phospho_site", "site_group_id",
            "heptameric_peptide", "pa_version",
        ])
        writer.writeheader()
        for e in all_entries:
            writer.writerow({k: e.get(k, "") for k in writer.fieldnames})
    print(f"  CSV export: {csv_path}")


if __name__ == "__main__":
    main()
